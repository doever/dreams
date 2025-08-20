# *-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-* *-*-*-*-*-*-*-* #
#                                 merge_data.py                                                  #
# *-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-* *-*-*-*-*-*-*-* #
# Description:                                                                                #
#     This script compares two files.                                                         #
#                                                                                             #
# Author     : cl                                                                             #
# Version    : v1.0                                                                           #
# CreTime    : 2025/4/14                                                                      #
# License    : Copyright (c) 2024 by cl, All rights reserved                                  #
# *-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-* #

import os
import re

from collections import defaultdict
from openpyxl import Workbook, load_workbook


def main():
    base_path = 'datas'
    files = os.listdir(base_path)
    print(files)
    groups = defaultdict(list)
    # 正则提取年份区间
    for name in files:
        match = re.search(r'(\d{2}-\d{2})', name)
        if match:
            year_range = match.group(1)
            groups[year_range].append(name)

    for year, files in groups.items():
        print(f"{year} 年份组：")
        temp_data = defaultdict(dict)
        all_fields = set()
        for file in files:
            # print(f"  - {file}")
            wb = load_workbook(os.path.join(base_path, file))
            ws = wb.active
            headers = list(ws.iter_rows(values_only=True))[0]
            rows = list(ws.iter_rows(values_only=True))[1:]
            for row in rows:
                seqn = str(row[0])
                for key, value in zip(headers[1:], row[1:]):
                    if key:
                        key = file[:-10] + '_' + key
                        temp_data[seqn][key] = value
                        all_fields.add(key)

        # 合并相同年份的多个excel，缺失值补none
        result = {}
        all_fields = sorted(all_fields)
        for seqn, info in temp_data.items():
            completed_info = {field: info.get(field, None) for field in all_fields}
            result[seqn] = completed_info

        # 写入到excel
        wb_ = Workbook()
        ws_ = wb_.active
        ws_.title = "MergedData"
        fieldnames = list(next(iter(result.values())).keys())
        print(fieldnames)
        header = ['seqn'] + fieldnames
        ws_.append(header)

        # 写入每一行数据
        for seqn, record in result.items():
            row = [seqn] + [record.get(field) for field in fieldnames]
            ws_.append(row)

        # 保存为 Excel 文件
        wb_.save(f"{year}.xlsx")


if __name__ == '__main__':
    main()
